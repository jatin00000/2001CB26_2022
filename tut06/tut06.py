#Import the necessary modules
import pandas
import openpyxl
from datetime import timedelta, date
from openpyxl import load_workbook
from datetime import datetime
 
#This line of Pandas->inputoutput->format->file_type = Excel-> ExcelFormatter is used to format excel way of looking
#Header_Style determines the header row look and all the layout related to header
#Setting it none make header row look likes normal data row only
#It is done on specific requirement of appearance of Output, as specified by Sir
pandas.io.formats.excel.ExcelFormatter.header_style = None

#As pandas raise settingWithCopyWarning,
#As it occur due to chain assignment,
#so to avoid its display in terminal, i used options attribute which is used to configure and prevent it from displaying error and exception
pandas.options.mode.chained_assignment = None

#During execution, Pandas was raising a lot of FutureWarnings
#to Prevent them from being displayed
#importing module 'warnings' which provide a lot a functionalities to handle warnings
#simplefilter() is use to apply a certain action on specific type of warnings
#applied action of 'ignore' to 'FutureWanings', 'UserWarning' category
import warnings
warnings.simplefilter(action='ignore', category= (FutureWarning, UserWarning))

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import csv
from random import randint
from time import sleep



def send_mail(fromaddr, frompasswd, toaddr, msg_subject, msg_body, file_path):
    try:
        msg = MIMEMultipart()
        print("[+] Message Object Created")
    except:
        print("[-] Error in Creating Message Object")
        return

    msg['From'] = fromaddr

    msg['To'] = toaddr

    msg['Subject'] = msg_subject

    body = msg_body

    msg.attach(MIMEText(body, 'plain'))

    filename = file_path
    attachment = open(filename, "rb")

    p = MIMEBase('application', 'octet-stream')

    p.set_payload((attachment).read())

    encoders.encode_base64(p)

    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)

    try:
        msg.attach(p)
        print("[+] File Attached")
    except:
        print("[-] Error in Attaching file")
        return

    try:
        #s = smtplib.SMTP('smtp.gmail.com', 587)
        s = smtplib.SMTP('mail.iitp.ac.in', 587)
        print("[+] SMTP Session Created")
    except:
        print("[-] Error in creating SMTP session")
        return

    s.starttls()

    try:
        s.login(fromaddr, frompasswd)
        print("[+] Login Successful")
    except:
        print("[-] Login Failed")

    text = msg.as_string()

    try:
        s.sendmail(fromaddr, toaddr, text)
        print("[+] Mail Sent successfully")
    except:
        print('[-] Mail not sent')

    s.quit()


def isEmail(x):
    if ('@' in x) and ('.' in x):
        return True
    else:
        return False

FROM_ADDR = "changeme"
FROM_PASSWD = "changeme"
receiver = "mayank265@iitp.ac.in"
Subject = "Sending Attendance Report"
Body ='''
Respected Sir,

Please find attached consolidated attendance report of students of CS384 python course with this mail.   

Thanking You.

--
Yours Sincerely
2001CB26
'''

from datetime import datetime
start_time = datetime.now()

def attendance_report():
    pointer1=pandas.read_csv("input_registered_students.csv")
    #name and roll no of students are stored in it
    rollno=[]
    names=[]
    # rollno=pointer1["Roll No"]
    for index in pointer1.index:
        roll=pointer1["Roll No"][index]
        rollno.append(roll)

    for index in pointer1.index:
        nam=pointer1["Name"][index]
        nam=rollno[index]+" "+nam
        names.append(nam)


    # names stored the roll no. and name
    try:
        
        attendance=pandas.read_csv("input_attendance.csv")
    except:
        
        print("input_attendence file not found")
        


    #finding total no_of rows and columns
    no_of_rows=len(attendance)
    no_of_col=len(attendance.columns)

    ##############################################
    #taking start time
    start=attendance["Timestamp"][0]
    end=attendance["Timestamp"][no_of_rows-1]




    start_time_stamp=pandas.Timestamp(start)
    startday=start_time_stamp.time()
    # start=start[0:10]
    # end=end[0:10]
    # it will stare date from start to end
    dateranges=pandas.date_range(start,end,freq='d')

    date_to_append_in=[]

    for i in range(len(dateranges)):
        date_mon_th=pandas.Timestamp(dateranges[i])
        

        proper_date=date_mon_th.date()
        proper_date= proper_date.strftime("%d-%m-%Y")

        date_day=date_mon_th.day_name()

        
        if(date_day=="Monday" or date_day=="Thursday"):
            date_to_append_in.append(proper_date)
    # print(date_to_append_in)
    # exit()        
    ##########################################################
    #opening work book named sheet 2
        
    lb=openpyxl.Workbook()
    pointer2=lb.active 
    pointer2.cell(row=1,column=1).value="Roll"
    pointer2.cell(row=1,column=2).value="Name"
    for  dd in range(len(date_to_append_in)):
        pointer2.cell(row=1,column=dd+3).value=date_to_append_in[dd]

    total_dates=len(date_to_append_in)
    pointer2.cell(row=1,column=total_dates+3).value="Actual Lecture Taken"
    pointer2.cell(row=1,column=total_dates+4).value="Total Real"
    pointer2.cell(row=1,column=total_dates+5).value=" %Attendance "

    #names basic information in sheet
    
    for i in range(len(names)): 
        
        fi=names[i]
        Filename_roll=fi[0:8]
        candidate_name=fi[8:]

        wb=openpyxl.Workbook()
        sheet=wb.active
        ##########################
        pointer2.cell(row=i+2,column=1).value=Filename_roll
        pointer2.cell(row=i+2,column=2).value=candidate_name
        
        
        
        ################
        date_dict={}
        for ii in range(len(date_to_append_in)):
            date_dict[date_to_append_in[ii]]=[0,0,0] 
            
            
        # print(date_dict) 
        ###############################################
    #   basic fillings in the sheet 
        sheet.cell(row=1,column=1).value="Date"
        sheet.cell(row=1,column=2).value="Roll"
        sheet.cell(row=1,column=3).value="Name"
        sheet.cell(row=1,column=4).value="Total Attendance Count"
        sheet.cell(row=1,column=5).value="Real"
        sheet.cell(row=1,column=6).value="Dublicate"
        sheet.cell(row=1,column=7).value="Invalid"
        sheet.cell(row=1,column=8).value="Absent"
        
        sheet.cell(row=2,column=2).value=Filename_roll
        sheet.cell(row=2,column=3).value=candidate_name 
    ################################################# 
        # datefillings
        
        for iii in range(len(date_to_append_in)):
            sheet.cell(row=iii+3,column=1).value=date_to_append_in[iii]
            


        for index in attendance.index:

            fullname=attendance["Attendance"][index]
            capitalroll=fullname[0:8]
            capitalname=fullname[8:]
            fullname=capitalroll+(capitalname.upper())
            #extracting name and roll num
            
            
            if(fullname==names[i]):
                # print(fullname)
                # print(fullname )
                # print(names[i])
                # print(" ")
                attendingtime=attendance["Timestamp"][index]
                my_date = datetime.strptime(attendingtime, "%d-%m-%Y %H:%M")
                day=my_date.weekday()
                hou=my_date.hour
                min=my_date.minute
                date_of_day= my_date.strftime("%d-%m-%Y")
                
                if(day==0 or day==3):
                    if( hou==14 or (hou==15 and min==0 )):
                        if(date_dict[date_of_day][0]>0):
                            date_dict[date_of_day][1]+=1
                        else:
                            date_dict[date_of_day][0]+=1
                    else:
                        date_dict[date_of_day][2]+=1

        no_of_A=0
        no_of_p=0
        # print(date_dict)
        # exit()
        #storing present absent and dublicate in date dict
        #  filling values in the sheet
        for xx in range(len(date_dict)):
            
            sheet.cell(row=3+xx,column=4).value=date_dict[date_to_append_in[xx]][0]+date_dict[date_to_append_in[xx]][1] +date_dict[date_to_append_in[xx]][2]  
            sheet.cell(row=3+xx,column=5).value=date_dict[date_to_append_in[xx]][0]        
            sheet.cell(row=3+xx,column=6).value=date_dict[date_to_append_in[xx]][1]
            sheet.cell(row=3+xx,column=7).value=date_dict[date_to_append_in[xx]][2]  
            real= date_dict[date_to_append_in[xx]][0]
            if(real):
                pointer2.cell(row=i+2,column=xx+3).value="P"
                no_of_p+=1
            else:
                pointer2.cell(row=i+2,column=xx+3).value="A"
                no_of_A+=1
                
            
            absent=1;
            if real==1:
                absent=0
            
            sheet.cell(row=3+xx,column=8).value=absent                

        #############################################
        pointer2.cell(row=i+2,column=total_dates+3).value=no_of_A+no_of_p
        pointer2.cell(row=i+2,column=total_dates+4).value=no_of_p
        
        trun=no_of_p/(no_of_A+no_of_p)
        
        trun=round(trun,2)
        pointer2.cell(row=i+2,column=total_dates+5).value=trun*100

        

        
        ##########################################
        date_dict.clear()
        #clearing the dictonary so that it can be used for other studen
        variablename=Filename_roll
        finalname="attendance_report_consolidate"
        lb.save(r"output/%s.xlsx"%finalname)
        wb.save( r"output/%s.xlsx"%variablename)
        
        #saving the file  in output file
                     

    
    

###Code

from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


attendance_report()
#what a ever is the limit of your sending mails, like gmail has 500.
max_count = 9999999
count=0
file_path = "output\attendance_report_consolidated.xlsx"
try:
        if isEmail(receiver):
            send_mail(FROM_ADDR, FROM_PASSWD, receiver, Subject, Body, file_path)
        print("Count Value: ", count)
        print("Sleeping . .. . ")
        sleep(randint(1,3))
except:
        print("Lets see ")





#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
