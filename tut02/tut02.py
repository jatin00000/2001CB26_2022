#Import the necessary modules
#I will be reading file through using openpyxl, pandas both
import pandas

#importing necessary things to make cell colorful
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.colors import Color
#As pandas raise settingWithCopyWarning,
#As it occur due to chain assignment,
#so to avoid its display in terminal, i used options attribute which is used to configure and prevent it from displaying error and exception
pandas.options.mode.chained_assignment = None

#Octant function takes three coordinates as parameter and return the octant to which these belong
def octant(x,y,z):
    if z>=0.000000000:
        if(x>=0.000000000 and y>=0.000000000  ): return 1
        elif(x<0.000000000  and y>=0.000000000 ): return 2
        elif(x<0.000000000  and y<0.000000000 ): return 3
        elif(x>=0.000000000 and y<0.000000000 ): return 4
    else:
        if(x>=0.000000000  and y>=0.000000000 ): return -1
        elif(x<0.000000000  and y>=0.000000000 ): return -2
        elif(x<0.000000000 and y<0.000000000 ): return -3
        elif(x>=0.000000000  and y<0.000000000 ): return -4
def Value_put(Pointer2, mod_range, counter, Pointer_Transition_Range, title):
    #Describing the parameters
    # Pointer2 is a file pointer
    # mod_range is a fstring containing in the format "<start value>-<end value>"
    # title is used to pass the following for printing 'Overall Count' or 'Mod Transition Count'
    # counter is a reference to print various row and column of matrix
    # Pointer_Transition_Range is a pointer to dictionary for particular transition range

    #1 Inserting the values which are fixed
    Pointer2['Octant ID'][counter]=title
    Pointer2['Octant ID'][counter+1]=mod_range
    Pointer2['1'][counter+1]='To'
    Pointer2[' '][counter+3]='From'
    Pointer2['Octant ID'][counter+2]='Count'
    Pointer2['Octant ID'][counter+3]='+1'
    Pointer2['Octant ID'][counter+4]='-1'
    Pointer2['Octant ID'][counter+5]='+2'
    Pointer2['Octant ID'][counter+6]='-2'
    Pointer2['Octant ID'][counter+7]='+3'
    Pointer2['Octant ID'][counter+8]='-3'
    Pointer2['Octant ID'][counter+9]='+4'
    Pointer2['Octant ID'][counter+10]='-4'
    Pointer2['1'][counter+2]='+1'
    Pointer2['-1'][counter+2]='-1'
    Pointer2['2'][counter+2]='+2'
    Pointer2['-2'][counter+2]='-2'
    Pointer2['3'][counter+2]='+3'
    Pointer2['-3'][counter+2]='-3'
    Pointer2['4'][counter+2]='+4'
    Pointer2['-4'][counter+2]='-4'

    #Space_according_to_Octant is a dictionary which contain position for particular transition value column as they need to come in a particular order
    Space_according_to_Octant ={'1':3, '-1':4,'2':5,'-2':6,'3':7,'-3':8,'4':9, '-4':10} #{octant: position along horizontal direction}
    #Feeding main values in the matrix
    for i in ['1','-1','2','-2','3','-3','4','-4']:
        for j in ['1','-1','2','-2','3','-3','4','-4']:
            Pointer2[i][counter+Space_according_to_Octant[j]]=Pointer_Transition_Range[f'{i}{j}']
            
def octant_transition_count(mod=5000):
    Pointer1 = pandas.read_excel("input_octant_transition_identify.xlsx")
    Pointer1.to_excel("output_octant_transition_identify.xlsx", index=False)
    Pointer2 = pandas.read_excel("output_octant_transition_identify.xlsx")
    #1counting the number of rows in csv file and computing the average of each U, V, W
    #shape attribute of pandas return number of rows and columns
    row__, column__ = Pointer2.shape #variable to keep count of number of rows

    avg_U = Pointer2['U'].sum()
    avg_V = Pointer2['V'].sum()
    avg_W = Pointer2['W'].sum()
    #computing average which is total sum of observations / number of observation
    avg_U= round(1.0*avg_U/row__,9) #round to 9 decimal places
    avg_V= round(1.0*avg_V/row__,9)
    avg_W= round(1.0*avg_W/row__,9)

    #2inserting using insert(position to insert, value of column, value by which every cell will be filled)
    #creating seven column and inserting blank value
    Pointer2.insert(len(Pointer2.columns), 'U Avg', '')
    Pointer2.insert(len(Pointer2.columns), 'V Avg', '')
    Pointer2.insert(len(Pointer2.columns), 'W Avg', '')
    Pointer2.insert(len(Pointer2.columns), 'U\'=U - U avg', '')
    Pointer2.insert(len(Pointer2.columns), 'V\'=V - V avg', '')
    Pointer2.insert(len(Pointer2.columns), 'W\'=W - W avg', '')
    Pointer2.insert(len(Pointer2.columns), 'Octant', '')

    #Rounding the column value upto 9 decimal values
    Pointer2['U Avg'][0]=round(avg_U,9)
    Pointer2['V Avg'][0]=round(avg_V,9)
    Pointer2['W Avg'][0]=round(avg_W,9)

    #creating a dictionary to keep count of the eight octant
    _hash = {'1':0, '-1':0, '2':0, '-2':0, '3':0, '-3':0, '4':0, '-4':0}
    
    #iterrows() a similar function as enumerate()
    for counter, rows in Pointer2.iterrows():
        #we can access any row of a particular column by
        #<file_pointer>['<column_label>'][counter] = <value>
        Pointer2['U\'=U - U avg'][counter]=('{:.9f}'.format(float(Pointer2['U'][counter])-avg_U)) #subtracting individual reading from average
        Pointer2['V\'=V - V avg'][counter]=('{:.9f}'.format(float(Pointer2['V'][counter])-avg_V))
        Pointer2['W\'=W - W avg'][counter]=('{:.9f}'.format(float(Pointer2['W'][counter])-avg_W))
        #calling the octant() function to give octant value, and storing it to the cell in octant colunn
        Pointer2['Octant'][counter]=octant(round(float(Pointer2['U\'=U - U avg'][counter]),9), round(float(Pointer2['V\'=V - V avg'][counter]),9), round(float(Pointer2['W\'=W - W avg'][counter]),9))
        #increasing the octant count in dictionary
        _hash[str(Pointer2['Octant'][counter])] +=1

    #3 Adding the remaining columns same as adding columns U avg, V avg
    Pointer2.insert(len(Pointer2.columns), ' ', '')
    Pointer2.insert(len(Pointer2.columns), 'Octant ID', '')
    Pointer2.insert(len(Pointer2.columns), '1', '')
    Pointer2.insert(len(Pointer2.columns), '-1', '')
    Pointer2.insert(len(Pointer2.columns), '2', '')
    Pointer2.insert(len(Pointer2.columns), '-2', '')
    Pointer2.insert(len(Pointer2.columns), '3', '')
    Pointer2.insert(len(Pointer2.columns), '-3', '')
    Pointer2.insert(len(Pointer2.columns), '4', '')
    Pointer2.insert(len(Pointer2.columns), '-4', '')

    #4 Filling the cells with values 'User Input', 'Overall count' as there position remain fixed whatever be the mod
    Pointer2[' '][1]='User Input'
    Pointer2['Octant ID'][1]='Mod '+str(mod)
    Pointer2['Octant ID'][0]='Overall Count'

    #5 Writing the overall count for each octant at respective position
    #There position will also remain fix, independent of value of mod
    Pointer2['1'][0]=_hash['1']
    Pointer2['-1'][0]=_hash['-1']
    Pointer2['2'][0]=_hash['2']
    Pointer2['-2'][0]=_hash['-2']
    Pointer2['3'][0]=_hash['3']
    Pointer2['-3'][0]=_hash['-3']
    Pointer2['4'][0]=_hash['4']
    Pointer2['-4'][0]=_hash['-4']

    #6 Generating the values that will be bound of intervals like 0,5000,10000 for mod = 5000 and storing them in list
    Bounds_mod_range=[]
    for count in range(0, row__, mod): #start with zero till number of rows and increment by mod
        Bounds_mod_range.append(count)
    Bounds_mod_range.append(row__+1)

    #Again making a dictionary, that will have for each octant all the intervals in form of list
    #Example {'1':[[0,4999,0], [5000,9999,0], .....], '-1':[[0,4999,0], [5000,9999,0], .....], ..........}
    #        'octant_value' : [[start_counter, end_counter, count of value]]
    Count_range_wise = {'1':[], '-1':[], '2':[], '-2':[], '3':[], '-3':[], '4':[], '-4':[]}
    for counter in range(len(Bounds_mod_range)-1):
        for count in Count_range_wise:
            Count_range_wise[count].append([Bounds_mod_range[counter],Bounds_mod_range[counter+1]-1,0])

    #7 For each mod range of each octant, count number of octant in it
    for counter, rows in Pointer2.iterrows():
        # print(counter)
        for idx in Count_range_wise[str(Pointer2['Octant'][counter])]:
                if counter>=idx[0] and counter<idx[1]:
                    #if lies in this interval, increment the count
                    idx[2]+=1
                    #we don't need to look furthur
                    break

    #8 Adding mod range labels in octant_output.csv file
    for count in range(len(Bounds_mod_range)-1):
        #counter+2 as they will start to fill from 2nd row in Octant ID column
        #will insert the range as an string
        Pointer2['Octant ID'][count+2]=str(Count_range_wise['1'][count][0])+'-'+str(Count_range_wise['1'][count][1])

    #9 Filling count of each octant in each mod range into csv file
    for key, value in Count_range_wise.items(): #iterating through dictionary
        for counter in range(len(Bounds_mod_range)-1): #number of ranges of mod will be constant
            Pointer2[str(key)][counter+2]=int(Count_range_wise[key][counter][2])
            #<file_handler>['<column_for_octant>'][row number] = string of Count_range_wise[key is octant][serial number of mod range][count of values]
    
    #10 Storing each transition as combination using fstring like 11 for +1 to +1 and 2-1 for +2 to -1 in a dictionary
    Transition_comb =dict()
    try :
        for i in ['1','-1','2','-2','3','-3','4','-4']:
            for j in ['1','-1','2','-2','3','-3','4','-4']:
                Transition_comb[f'{i}{j}']=0 #here key is fstring and value is count of that key
    except ValueError():
        print("ValueError in Part 10")
    except :
        print("Other error in part 10")

    #11 making another dictionary such that for each transition range
    # starting bound is the key and value is the above dictionary which is Transition_comb for each range
    Transition_range_comb = dict()
    for i in range(len(Bounds_mod_range)-1):
        val = Bounds_mod_range[i]
        Transition_range_comb[val] = Transition_comb.copy()

    #12 For each counter, we will count transition
    try: 
        for counter, rows in Pointer2.iterrows():
            if counter==0: 
                continue #skip first counter as there is no row above it to make a transition
            else :
                val = mod * int(counter/mod) 
                #it is formula to find lower bound of a range to which a counter belongs
                # for 11555, counter/mod = 11555/5000 = 2.311
                # int(2.311) = 2
                # mod *2 = 5000*2 = 10,000 
                i = Pointer2['Octant'][counter-1]
                j = Pointer2['Octant'][counter]
                #using fstring, make suitable key
                Transition_range_comb[val][f'{i}{j}'] += 1
                #increasing Transition_comb[key] to keep transition count for overall
                Transition_comb[f'{i}{j}'] +=1
    except ValueError():
        print("ValueError in Part 12")
    except:
        print("Error in Part 12")


    #variable to keep track of position to print Matrix of various ranges of transition count
    counter = len(Bounds_mod_range)+1
    #13 Writing 'Verified' at its position, its position remain fixed
    Pointer2['Octant ID'][counter]='Verified'

    try: 
        #inserting verified values for each octant
        Pointer2['1'][counter]=_hash['1']
        Pointer2['-1'][counter]=_hash['-1']
        Pointer2['2'][counter]=_hash['2']
        Pointer2['-2'][counter]=_hash['-2']
        Pointer2['3'][counter]=_hash['3']
        Pointer2['-3'][counter]=_hash['-3']
        Pointer2['4'][counter]=_hash['4']
        Pointer2['-4'][counter]=_hash['-4']
    except ValueError():
        print("ValueError in Part 13")
    except:
        print("Error in Part 13")

    #leaving difference of 3 row to write Overall transition count
    counter+=3
    #Writing overall transition count table using Value_put function
    Value_put(Pointer2, ' ', counter, Transition_comb, 'Overall Transition Count')

    #Leaving space of rows, 
    #counter = 1 row for 'To' + 1 row for 'From' + 8 rows for 8 octants + 2 row for blank space + 1 for beginning of next matrix
    counter+=14

    try: 
        #14 Using For loop, putting each matrix in excel file
        for i in range(len(Bounds_mod_range)-1): #For each range
            val = Bounds_mod_range[i]
            #Calling Function Value_put()
            Value_put(Pointer2, f'{Bounds_mod_range[i]}-{Bounds_mod_range[i+1]-1}', counter, Transition_range_comb[val], 'Mod Transition Count')
            counter+=13 #Jumping to next desired location
    except TypeError():
        print("TypeError in Part 14")
    except ValueError():
        print("ValueError in Part 14")
    except:
        print("Error in Part 14")
    
    #15 Earlier, values in this column was 1 instead of +1
    #So Making it with sign using format specifier
    Pointer2['Octant']  = Pointer2['Octant'].apply(lambda x: '{:+d}'.format(x))

    #16 Renaming these columns according to this specification
    Pointer2.rename(columns = {'Octant ID':' ', 'Octant':'Octact', '1':'+1', '2':'+2', '3':'+3', '4':'+4'}, inplace = True)

    #17 Saving all changes to output file
    Pointer2.to_excel("output_octant_transition_identify.xlsx", index=False)

    #18 Again opening file using openpyxl
    wb = load_workbook("output_octant_transition_identify.xlsx")
    ws = wb["Sheet1"] #Making current sheet active

    #Coloring the cell with 'User Input' according to specification
    ws["L3"].fill = PatternFill("solid", start_color="FFFF00")
    #PatternFill function is used to give background color to cell

    #Formatting the cell with value 'Mod 5000' according to specification
    ws["M3"].fill = PatternFill("solid", start_color="C6EFCE")
    #Font() is used to give font color to the cell
    ws["M3"].font =  Font(color = "006100")

    #19 Saving after adding color
    wb.save("output_octant_transition_identify.xlsx")

###Code

from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

mod=5000
octant_transition_count(mod)
